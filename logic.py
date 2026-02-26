import pandas as pd
import pyodbc
import re
import copy
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ------------------------------------------------------------
# 1. CONFIGURACIÓN DB
# ------------------------------------------------------------
DB_IP = os.getenv("DB_IP", "192.168.0.253")
DB_USER = os.getenv("DB_USER", "debo")
DB_PASS = os.getenv("DB_PASS", "debo")
DB_NAME = os.getenv("DB_NAME", "DEBO")

PRECIO_GNC_DEFAULT = 669

RESPONSABLES_GNC_VALIDOS = [
    "RICARDO RAMON",
    "KARINA ADRIANA",
    "HEBER AGUSTIN",
    "RAMON ELECTO",
    "JAVIER ANTONIO",
    "MIGUEL ALEJANDRO",
    "MATIAS CASAS",
    "PONCE SANTIAGO",
    "ESTEBAN ORTIZ",
    "MARLENE GONZALEZ",
    "ARCE RUBEN EMANUEL",
    "SABRINA RUIZ",
    "ARGUELLO MARCOS",
    "ELIAS GARCIA",
    "LUCAS ROMERO",
    "NICOLAS SUAREZ",
    "ROMAN VELAZQUEZ",
    "BRIAN VEGA",
    "RODRIGO BREZESKY",
    "LUCAS MIRAGLIA",
]

# -------------------------
# ESTADOS QR (GLOBAL)
# -------------------------
ESTADO_QR_PENDIENTE = "PENDIENTE"
ESTADO_QR_IMPACTADO = "IMPACTO_QR"
ESTADO_QR_SIN_COMPROBANTE = "SIN_COMPROBANTE"
ESTADO_QR_MANUAL = "ASIGNADO_MANUAL"
ESTADO_QR_ASIGNADO_MANUAL = "ASIGNADO_MANUAL"


class AforadorGNC:
    def __init__(self, inicial=0.0, final=0.0, precio=PRECIO_GNC_DEFAULT):
        self.inicial = float(inicial or 0)
        self.final = float(final or 0)
        self.precio = float(precio or PRECIO_GNC_DEFAULT)

    def consumo(self):
        return max(self.final - self.inicial, 0)

    def total(self):
        return self.consumo() * self.precio


class CoberturaGNC:
    """
    Usado para:
    - Cubrió GNC
    - Baños
    (misma lógica)
    """
    def __init__(self, nombre, responsable, aforadores, vendedor_cubre=None):
        self.nombre = nombre
        self.responsable = responsable
        self.aforadores = aforadores  # lista de AforadorGNC
        self.vendedor_cubre = vendedor_cubre

    def total(self):
        return sum(a.total() for a in self.aforadores)


def calcular_gnc_general(aforadores):
    """
    GNC total del turno (NO afectado por coberturas)
    """
    return sum(a.total() for a in aforadores)


def obtener_conexion_sql():
    """Intenta conectar a la base de datos SQL Server y maneja errores."""
    try:
        conn_str = (
            f"DRIVER={{SQL Server}};SERVER={DB_IP};DATABASE={DB_NAME};"
            f"UID={DB_USER};PWD={DB_PASS};Connection Timeout=10;"
        )
        return pyodbc.connect(conn_str)
    except Exception as e:
        print(f"Error de Red: No se pudo conectar a la DB:\n{e}")
        return None

def normalizar_texto(texto):
    """Limpia el texto, elimina prefijos numerados y lo convierte a mayúsculas."""
    if not texto or str(texto).lower() == 'nan':
        return ""
    # Elimina patrones como "1 - " o "23-"
    return re.sub(r'^\d+\s*-\s*', '', str(texto)).strip().upper()


def parse_moneda_robusto(valor):
    """Convierte un valor de texto o numérico (con manejo de comas/puntos) a float."""
    try:
        if valor is None or str(valor).lower() == 'nan' or valor == '':
            return 0.0
        if isinstance(valor, (float, int)):
            return float(valor)

        s = str(valor).strip()
        # Caso 1.234,56 (formato europeo con separador de miles y coma decimal)
        if ',' in s and '.' in s:
            s = s.replace(".", "").replace(",", ".")
        # Caso 1234,56 (formato con coma decimal)
        elif ',' in s:
            s = s.replace(",", ".")

        return float(s)
    except:
        return 0.0


def formatear_arg(valor):
    """Formatea un número a formato de moneda ARS (1.234,56) e incluye signo."""
    prefix = "+" if valor > 0.001 else ""
    # "{:,.2f}" genera 1,234.56, luego se invierten , y .
    val_str = "{:,.2f}".format(valor).replace(",", "v").replace(".", ",").replace("v", ".")
    return f"{prefix}{val_str}"


def son_nombres_similares(excel, db):
    ex, db = normalizar_texto(excel), normalizar_texto(db)
    if not ex or not db:
        return False

    pal_ex = set(ex.split())
    pal_db = set(db.split())

    # 🔧 FORZAR: si DB es subconjunto del Excel → MATCH
    if pal_db.issubset(pal_ex):
        return True

    # reglas manuales existentes
    if "ROMAN" in db or "ROMAN" in ex:
        if "VELAZQUEZ" in db and "VELAZQUEZ" in ex:
            return True
        return False

    return pal_ex.issubset(pal_db)


def normalizar_desc_promo(texto):
    if not texto:
        return 0.0

    s = str(texto)

    # Captura números con miles y decimales en ambos formatos
    nums = re.findall(r"\d[\d.,]*", s)
    if not nums:
        return 0.0

    total = 0.0

    for n in nums:
        original = n

        # 🇦🇷 Formato argentino: 4.800,00
        if "." in n and "," in n and n.rfind(".") < n.rfind(","):
            n = n.replace(".", "").replace(",", ".")
        # 🇺🇸 Formato USA: 4,800.00
        elif "." in n and "," in n and n.rfind(",") < n.rfind("."):
            n = n.replace(",", "")
        # Miles sin decimales: 4.800 o 4,800
        elif n.count(".") == 1 and len(n.split(".")[1]) == 3:
            n = n.replace(".", "")
        elif n.count(",") == 1 and len(n.split(",")[1]) == 3:
            n = n.replace(",", "")
        # Decimal con coma
        elif "," in n:
            n = n.replace(",", ".")

        try:
            total += float(n)
        except ValueError:
            pass

    return round(total, 2)


def limpiar_texto_monetario(valor):
    """
    Extrae el número (con signo) de un texto.
    Ejemplos:
    '$1.200 promo' -> '1.200'
    'promo -500'   -> '-500'
    'DESC $300'    -> '300'
    """
    if valor is None:
        return ""
    s = str(valor)
    # deja solo dígitos, coma, punto y signo -
    limpio = re.findall(r'-?\d+[.,]?\d*', s)
    return limpio[0] if limpio else ""

def normalizar_id_transaccion(val):
    if val is None:
        return ""
    s = str(val)
    try:
        if s.lower().startswith("0x"):
            return str(int(s, 16))
    except:
        pass
    return s

def obtener_turno_actual():
    hora = datetime.now().hour

    if 6 <= hora < 15:
        return "Turno_Mañana"
    elif 15 <= hora < 23:
        return "Turno_Tarde"
    else:
        return None  # no se guarda

def aplicar_header(ws, headers):
    header_fill = PatternFill("solid", fgColor="2c3e50")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    ws.append(headers)

    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center")
        c.border = border

    ws.freeze_panes = "A2"

def ajustar_columnas(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 40)

class AuditManager:
    def __init__(self):
        # State variables
        self.planilla_desde_sql = None
        self.planilla_hasta_sql = None

        self.datos_detalle_qr = {}
        self.datos_detalle_facturacion = {}
        self.datos_rendiciones = {}
        self.df_rendiciones_cache = None
        self.df_gnc_sql_cache = None
        self.df_vendedores_cache = None

        self.anotaciones_tmp = []
        self.qr_reasignados = set()

        self.gnc_general_total = 0.0
        self.gnc_coberturas = []
        self.gnc_total_coberturas = 0.0
        self.gnc_para_caja = 0.0
        self.gnc_extra_por_responsable = {}

        self.cajas_data = {}
        self.descuentos_qr_por_vendedor = {}

        self.turno_seleccionado = None
        self.transaccion_qr_buscada = None

        # Vendor Input State (Replaces UI widgets)
        # Structure: vendor_name -> { 'base': float, 'prod': float, 'per': float,
        #                             'tarj': float, 'ef': float, 'caja_id': set,
        #                             'fusionado': bool, 'gnc_asignado': float, 'gnc_base': float }
        self.vendor_states = {}

    def fetch_sql_data(self, desde, hasta):
        self.planilla_desde_sql = desde
        self.planilla_hasta_sql = hasta
        self.datos_detalle_facturacion = {}

        conn = obtener_conexion_sql()
        if not conn:
            raise Exception("No se pudo conectar a la base de datos.")

        try:
            sql_qr = """
            ;WITH QR_BASE AS (
                SELECT
                    A.PLA AS Planilla,
                    COALESCE(v2.NOMVEN, v1.NOMVEN, 'SIN OPERARIO') AS Operario,
                    A.NTRANS AS ID_TRANSACCION,
                    CAST(REPLACE(A.IMPORTE, ',', '.') AS DECIMAL(18,2)) AS IMPORTE,
                    CAST(REPLACE(ISNULL(A.CASHOUT, '0'), ',', '.') AS DECIMAL(18,2)) AS CASHOUT,
                    A.RED_DES AS DESC_PROMO,
                    CAST(REPLACE(ISNULL(A.RED_TOT,'0'), ',', '.') AS DECIMAL(18,2)) AS DESCUENTO_TOTAL,
                    A.FEC, A.ORI, A.LUG, A.EST AS EST_MP,
                    A.CAR AS ID_COMPV2,
                    af.TIP, af.TCO, af.NCO
                FROM A_MERCADOPAGO A
                LEFT JOIN AMAEFACT_EXT C ON A.CAR = C.ID_COMPV2
                LEFT JOIN AMAEFACT af ON af.SUC = C.SUC AND af.NCO = C.NCO AND af.TIP = C.TIP AND af.TCO = C.TCO
                LEFT JOIN VENDEDORES v1 ON af.VEN = v1.CODVEN
                LEFT JOIN VENDEDORES v2 ON af.OPE = v2.CODVEN
                WHERE A.PLA BETWEEN ? AND ? AND A.LUG = 1 AND A.EST = 1
            ),
            QR_DEPURADO AS (
                SELECT *,
                       ROW_NUMBER() OVER (PARTITION BY ID_COMPV2 ORDER BY NCO DESC) AS rn
                FROM QR_BASE
            ),
            QR_CLASIFICADO AS (
                SELECT *,
                    CASE
                        WHEN ID_COMPV2 IS NULL THEN 'QR SIN COMPROBANTE'
                        WHEN NCO IS NULL THEN 'QR COMPROBANTE INEXISTENTE'
                        ELSE 'QR OK'
                    END AS ESTADO_QR,
                    ROUND(IMPORTE + CASHOUT - DESCUENTO_TOTAL, 2) AS QR_NETO_FINAL
                FROM QR_DEPURADO
                WHERE rn = 1
            )
            SELECT
                Planilla, Operario, ESTADO_QR, ID_TRANSACCION, IMPORTE, CASHOUT,
                DESC_PROMO, DESCUENTO_TOTAL, QR_NETO_FINAL, FEC, TIP, TCO, NCO
            FROM QR_CLASIFICADO
            ORDER BY Operario, FEC;
            """

            sql_facturas = """
            SELECT
                f.FEC AS Fecha, f.TCO AS Tipo, f.NCO AS Numero, f.CPA AS Ref_CPA,
                COALESCE(v_origen.NomVen, v_actual.NomVen, 'DESCONOCIDO') AS Vendedor_Nombre,
                art.DetArt AS Producto,
                CAST(d.CAN AS DECIMAL(18,2)) AS Cantidad,
                CASE
                    WHEN f.TCO LIKE '%NC%' THEN (CAST(d.CAN AS DECIMAL(18,2)) * -1) * CAST(art.PreVen AS DECIMAL(18,2))
                    ELSE CAST(d.CAN AS DECIMAL(18,2)) * CAST(art.PreVen AS DECIMAL(18,2))
                END AS Total_Neto
            FROM AMAEFACT f
            LEFT JOIN AMAEFACT f_origen ON f.SUC = f_origen.SUC AND f.TCO LIKE '%NC%' AND f_origen.TCO NOT LIKE '%NC%' AND TRY_CAST(f.CPA AS BIGINT) = f_origen.NCO
            LEFT JOIN VENDEDORES v_origen ON f_origen.OPE = v_origen.CodVen
            LEFT JOIN VENDEDORES v_actual ON f.OPE = v_actual.CodVen
            INNER JOIN AMOVSTOC d ON f.SUC = d.PVE AND f.NCO = d.NCO AND f.TIP = d.TIP AND f.TCO = d.TCO
            INNER JOIN ARTICULOS art ON d.ART = art.CodArt
            WHERE f.PLA BETWEEN ? AND ? AND f.ANU = ''
            AND ((art.CodSec = 0 AND art.CodRub = 6) OR art.DetArt LIKE '%PELOTA%' OR (art.CodRub = 10 AND d.ART IN (7, 8, 9)))
            ORDER BY f.FEC, f.NCO;
            """

            sql_percepcion = """
            SELECT
                f.TCO,
                CAST(f.SUC AS VARCHAR(5)) + '-' + CAST(f.NCO AS VARCHAR(20)) AS Comprobante,
                f.TOT AS Total_Factura,
                f.PER AS Percepcion,
                v.NomVen AS Nombre_Vendedor
            FROM AMAEFACT f
            LEFT JOIN VENDEDORES v ON f.OPE = v.CodVen
            WHERE f.PLA BETWEEN ? AND ? AND f.TCO LIKE 'F%' AND f.TIP = 'A' AND f.ANU = ''
            ORDER BY f.SUC, f.NCO
            """

            sql_efectivo = """
            SELECT
                r.PLA AS Planilla, r.NUM AS Nro_Mov, r.OPE AS Operario,
                v.NomVen AS Vendedor, r.FEC AS Fecha, r.EFE AS Efectivo, r.LUG AS Sector
            FROM ATURRPA r
            LEFT JOIN VENDEDORES v ON r.OPE = v.CodVen
            WHERE r.PLA BETWEEN ? AND ? AND r.LUG = 1
            ORDER BY v.NomVen, r.FEC
            """

            self.df_rendiciones_cache = pd.read_sql(sql_efectivo, conn, params=[desde, hasta])
            self.df_gnc_sql_cache = self.df_rendiciones_cache.copy()

            df_qr = pd.read_sql(sql_qr, conn, params=(desde, hasta))

            # Normalize QR Data
            df_qr["IMPORTE"] = pd.to_numeric(df_qr["IMPORTE"], errors="coerce").fillna(0.0)
            df_qr["CASHOUT"] = pd.to_numeric(df_qr["CASHOUT"], errors="coerce").fillna(0.0)
            df_qr["DESC_PROMO_NUM"] = df_qr["DESC_PROMO"].apply(normalizar_desc_promo)

            df_qr = (df_qr.groupby("ID_TRANSACCION", as_index=False).agg({
                "Operario": "first", "IMPORTE": "first", "CASHOUT": "first",
                "DESC_PROMO": "first", "DESC_PROMO_NUM": "sum", "FEC": "first",
                "ESTADO_QR": "first", "TIP": "first", "TCO": "first", "NCO": "first",
            }))

            df_qr["QR_FINAL"] = (df_qr["IMPORTE"] + df_qr["CASHOUT"] - df_qr["DESC_PROMO_NUM"]).round(2)

            # Process Rendiciones
            # Clear previous SQL rendiciones
            for vend in self.datos_rendiciones:
                if 'movimientos' in self.datos_rendiciones[vend]:
                    self.datos_rendiciones[vend]['movimientos'] = [
                        m for m in self.datos_rendiciones[vend]['movimientos']
                        if m.get('origen') != 'sql'
                    ]

            # Load SQL Rendiciones
            for _, row in self.df_rendiciones_cache.iterrows():
                vendedor_sql = str(row['Vendedor']).strip().upper()
                monto = parse_moneda_robusto(row['Efectivo'])

                # Note: Logic must handle linking to existing vendors or creating new ones in state
                self.datos_rendiciones.setdefault(vendedor_sql, {}).setdefault('movimientos', []).append({
                    'origen': 'sql',
                    'planilla': row['Planilla'],
                    'nro': row['Nro_Mov'],
                    'tipo': 'Efectivo',
                    'ref': f"Mov {row['Nro_Mov']} - {row['Fecha']:%d/%m %H:%M}",
                    'monto': float(monto)
                })

            # Process QR per vendor
            self.datos_detalle_qr.clear()
            for operario in df_qr['Operario'].dropna().unique():
                df_qr_vendedor = df_qr[df_qr['Operario'] == operario].copy()
                df_qr_vendedor = df_qr_vendedor[df_qr_vendedor["ESTADO_QR"].isin([
                    "QR OK", "QR SIN COMPROBANTE", "QR COMPROBANTE INEXISTENTE"
                ])]
                if not df_qr_vendedor.empty:
                    df_qr_vendedor = df_qr_vendedor.drop_duplicates(subset=["ID_TRANSACCION"], keep="first")
                    self.datos_detalle_qr[operario] = df_qr_vendedor

            self.descuentos_qr_por_vendedor = {}
            for operario, df in self.datos_detalle_qr.items():
                self.descuentos_qr_por_vendedor[operario] = df["DESC_PROMO_NUM"].sum()

            # Facturas
            df_fact = pd.read_sql(sql_facturas, conn, params=(desde, hasta))
            df_per = pd.read_sql(sql_percepcion, conn, params=(desde, hasta))

            df_per_sum = df_per.groupby('Nombre_Vendedor', dropna=False)['Percepcion'].sum().reset_index()

            # Purge NC
            df_nc = df_fact[df_fact['Tipo'].str.contains('NC', na=False)].copy()
            numeros_anulados = set()
            for _, row in df_nc.iterrows():
                try:
                    ref = row['Ref_CPA']
                    if pd.notna(ref):
                        numeros_anulados.add(int(str(ref).strip()))
                except:
                    pass

            indices_a_borrar = []
            for idx, row in df_fact.iterrows():
                es_nc = 'NC' in str(row['Tipo'])
                try:
                    numero = int(str(row['Numero']).strip())
                except:
                    numero = -1
                if es_nc or numero in numeros_anulados:
                    indices_a_borrar.append(idx)

            df_fact = df_fact.drop(indices_a_borrar)

            # Store Fact Data
            for v_sql in df_fact['Vendedor_Nombre'].unique():
                self.datos_detalle_facturacion[v_sql] = df_fact[df_fact['Vendedor_Nombre'] == v_sql]

            # Update Vendor States with SQL Data
            # QR
            for vendedor_sql, df in self.datos_detalle_qr.items():
                total_qr = round(df["QR_FINAL"].sum(), 2)
                self.update_vendor_state(vendedor_sql, qr_db=total_qr)

            # Percepcion
            for _, r in df_per_sum.iterrows():
                nom_sql = r['Nombre_Vendedor']
                monto = parse_moneda_robusto(r['Percepcion'])
                self.update_vendor_state(nom_sql, per=monto)

            # Prod
            df_sumas = df_fact.groupby('Vendedor_Nombre')['Total_Neto'].sum().reset_index()
            for _, r in df_sumas.iterrows():
                nom_sql = r['Vendedor_Nombre']
                monto = r['Total_Neto']
                self.update_vendor_state(nom_sql, prod=monto)

        finally:
            conn.close()


    def update_vendor_state(self, vendor_name, **kwargs):
        """Updates or initializes the state for a specific vendor."""
        vendor_name = normalizar_texto(vendor_name)
        if vendor_name not in self.vendor_states:
            self.vendor_states[vendor_name] = {
                'base': 0.0, 'base_excel': 0.0, 'base_original': 0.0,
                'qr_db': 0.0, 'prod': 0.0, 'per': 0.0, 'tarj': 0.0, 'ef': 0.0,
                'gnc_asignado': 0.0, 'gnc_base': 0.0, 'gnc_aplicado': False,
                'caja_id': frozenset([vendor_name]), 'fusionado': False
            }

        state = self.vendor_states[vendor_name]
        for key, value in kwargs.items():
            if key in state:
                if key == 'caja_id' and not isinstance(value, frozenset):
                     state[key] = frozenset(value)
                else:
                     state[key] = value

    def load_vendedores_from_db(self):
        conn = obtener_conexion_sql()
        if not conn:
            return []
        try:
            query = "SELECT CodVen, NomVen FROM VENDEDORES"
            df = pd.read_sql(query, conn)
            return sorted(df["NomVen"].dropna().str.strip().str.upper().tolist())
        finally:
            conn.close()

    def calculate_balances(self):
        """
        Calculates the financial balance for each 'caja' (group of vendors).
        Returns a dictionary mapping vendor names to their difference/balance.
        """
        cajas = {}
        # Group vendors by caja_id
        for v, state in self.vendor_states.items():
            cid = state["caja_id"]
            cajas.setdefault(cid, []).append((v, state))

        results = {}
        total_caja_conjunta = 0.0

        for cid, lista in cajas.items():
            entrada_bruta = 0.0
            salida_real = 0.0
            ajustes = 0.0

            # Identify main vendor (not fusionado)
            fila_principal = None
            for v, state in lista:
                if not state.get("fusionado"):
                    fila_principal = v

                # Entrada
                base = state["base"] # Includes Excel + GNC + Coberturas
                prod = parse_moneda_robusto(state["prod"])
                per = parse_moneda_robusto(state["per"])
                entrada_bruta += base + prod + per

                # Salida
                qr = state.get("qr_db", 0.0)
                tarj = parse_moneda_robusto(state["tarj"])

                # Efectivo (from rendiciones)
                movs = self.datos_rendiciones.get(v, {}).get("movimientos", [])
                efectivo = sum(
                    parse_moneda_robusto(m["monto"])
                    for m in movs
                    if m.get("origen") in ("sql", "manual")
                )
                salida_real += qr + efectivo + tarj

                # Ajustes
                ajustes += self.total_anotaciones_por_vendedor(v)

                # Promo QR logic
                promo_qr = 0.0
                for vend_sql, total_promo in self.descuentos_qr_por_vendedor.items():
                    if son_nombres_similares(v, vend_sql):
                        promo_qr = total_promo
                        break
                ajustes += promo_qr

            entrada_real = entrada_bruta - ajustes
            diferencia = salida_real - entrada_real

            # Store result for all vendors in this caja
            for v, state in lista:
                results[v] = diferencia

            if fila_principal:
                 total_caja_conjunta += diferencia

        return results, total_caja_conjunta

    def total_anotaciones_por_vendedor(self, vendedor):
        total = 0.0
        for a in self.anotaciones_tmp:
            if a.get("estado") == ESTADO_QR_IMPACTADO:
                continue
            if son_nombres_similares(vendedor, a["vendedor"]):
                total += float(a["monto"])
        return total

    def add_anotacion(self, transaccion, vendedor, monto, descripcion, tipo="QR"):
        vend_norm = normalizar_texto(vendedor)
        self.anotaciones_tmp.append({
            "transaccion": str(transaccion).strip(),
            "vendedor": vend_norm,
            "monto": float(monto),
            "descripcion": descripcion.strip(),
            "tipo": tipo,
            "estado": ESTADO_QR_PENDIENTE,
            "origen_qr": ESTADO_QR_SIN_COMPROBANTE,
            "qr_fila": None
        })
        self.reconcile_anotaciones()

    def remove_anotacion(self, transaccion_id):
        self.anotaciones_tmp = [
            a for a in self.anotaciones_tmp
            if str(a.get("transaccion")) != str(transaccion_id)
        ]
        if str(transaccion_id) in self.qr_reasignados:
            self.qr_reasignados.remove(str(transaccion_id))
            # Also remove from rendiciones if it was impacted there?
            # Original code logic for removing rendicion is inside "eliminar_transaccion_seleccionada"
            # but mainly it just refreshes UI.
            # The apply_qr... function added to DATOS_RENDICIONES. We might need to clean that up.
            # For now, following original logic: just remove from list and refresh.
            pass

    def reconcile_anotaciones(self):
        if not self.datos_detalle_qr:
            return False

        hubo_cambios = False
        for anot in self.anotaciones_tmp:
            if anot.get("estado") != ESTADO_QR_PENDIENTE:
                continue

            nro = str(anot.get("transaccion", "")).strip()
            if not nro: continue

            estado, vendedor_qr, fila = self.search_qr_transaction(nro)

            if estado != "OK" or fila is None or not vendedor_qr:
                continue

            anot["estado"] = ESTADO_QR_IMPACTADO
            anot["qr_fila"] = fila

            # Logic from aplicar_qr_a_vendedor_desde_anotacion
            self._apply_qr_impact(anot)
            hubo_cambios = True

        return hubo_cambios

    def _apply_qr_impact(self, anot):
        nro = str(anot.get("transaccion"))
        if nro in self.qr_reasignados:
            return

        fila = anot["qr_fila"]
        monto = float(fila.get("QR_FINAL", 0.0))
        vendedor = anot["vendedor"]

        if monto == 0:
            return

        self.datos_rendiciones.setdefault(vendedor, {}).setdefault("movimientos", []).append({
            "origen": "qr_anotacion",
            "tipo": "QR reasignado",
            "ref": f"QR {nro}",
            "monto": monto
        })
        self.qr_reasignados.add(nro)

    def search_qr_transaction(self, id_transaccion):
        # Check if manually assigned (excluded from search)
        for a in self.anotaciones_tmp:
            if (str(a.get("transaccion")) == str(id_transaccion)
                and a.get("estado") == ESTADO_QR_ASIGNADO_MANUAL):
                return "NO_ENCONTRADA", None, None

        if not id_transaccion:
             return "NO_ENCONTRADA", None, None

        id_transaccion = str(id_transaccion).strip()

        for vendedor, df in self.datos_detalle_qr.items():
            if df is None or df.empty: continue
            if "ID_TRANSACCION" not in df.columns: continue

            coincidencia = df[df["ID_TRANSACCION"].astype(str) == id_transaccion]
            if coincidencia.empty: continue

            fila = coincidencia.iloc[0]
            comprobante = str(fila.get("NCO", "")).strip()
            vendedor_qr = fila.get("VENDEDOR") or fila.get("Operario") or vendedor

            if not comprobante or comprobante in ("", "None", "0"):
                 return "SIN_COMPROBANTE", fila, None

            return "OK", vendedor_qr, fila

        return "NO_ENCONTRADA", None, None

    def calculate_gnc(self, aforadores, coberturas, precio_gnc):
        self.gnc_general_total = sum(a.total() for a in aforadores)
        self.gnc_coberturas = coberturas
        self.gnc_total_coberturas = sum(c.total() for c in coberturas)
        self.gnc_para_caja = self.gnc_general_total - self.gnc_total_coberturas

        self.gnc_extra_por_responsable = {}
        for c in coberturas:
            self.gnc_extra_por_responsable.setdefault(c.responsable, 0.0)
            self.gnc_extra_por_responsable[c.responsable] += c.total()

        # Apply logic to update bases
        # Reset gnc_asignado
        for state in self.vendor_states.values():
            state["gnc_asignado"] = 0.0

        for c in self.gnc_coberturas:
            responsable = normalizar_texto(c.responsable)
            vendedor_cubre = normalizar_texto(c.vendedor_cubre)

            key_responsable = None
            key_cubre = None

            for vend in self.vendor_states.keys():
                if son_nombres_similares(vend, responsable): key_responsable = vend
                if son_nombres_similares(vend, vendedor_cubre): key_cubre = vend

            monto = c.total()
            if key_responsable: self.vendor_states[key_responsable]["gnc_asignado"] -= monto
            if key_cubre: self.vendor_states[key_cubre]["gnc_asignado"] += monto

        # Recalculate base
        for state in self.vendor_states.values():
            state["base"] = state["base_excel"] + state.get("gnc_base", 0.0) + state["gnc_asignado"]


    def generate_excel_report(self, turno, output_folder=None):
        if output_folder is None:
            output_folder = os.getenv("OUTPUT_FOLDER", "C:/cierres de caja/")
        if not turno:
            raise ValueError("Debe seleccionar un turno.")

        os.makedirs(output_folder, exist_ok=True)
        fecha = datetime.now().strftime("%Y-%m-%d")
        ruta = os.path.join(output_folder, f"Cierre_Caja_{turno}_{fecha}.xlsx")

        wb = Workbook()
        ws_resumen = wb.active
        ws_resumen.title = "RENDICIÓN FINAL POR VENDEDOR"

        headers = ["Planilla", "Vendedor", "Diferencia", "Observaciones"]
        aplicar_header(ws_resumen, headers)

        balances, _ = self.calculate_balances()

        for v, diff in balances.items():
            obs_desc = []
            for a in self.anotaciones_tmp:
                if not son_nombres_similares(v, a.get("vendedor")): continue
                if a.get("monto") in (None, "", 0): continue
                desc = (a.get("descripcion") or "").strip()
                if desc: obs_desc.append(desc)

            ws_resumen.append([
                self.planilla_desde_sql or "",
                v,
                diff,
                "; ".join(obs_desc)
            ])

            fila = ws_resumen.max_row
            color = "c0392b" if diff < 0 else "27ae60"
            ws_resumen[f"C{fila}"].font = Font(color=color, bold=True)
            ws_resumen[f"C{fila}"].number_format = '"$"#,##0.00;-"$"#,##0.00'

        ajustar_columnas(ws_resumen)

        # Details per vendor
        for v, state in self.vendor_states.items():
            ws = wb.create_sheet(v[:31])
            ws["A1"] = "Fecha:"
            ws["B1"] = fecha
            ws["A2"] = "Turno:"
            ws["B2"] = turno
            ws["A3"] = "Planilla:"
            ws["B3"] = self.planilla_desde_sql or ""

            ws["A5"] = "Concepto"
            ws["B5"] = "Importe"
            ws["A5"].font = ws["B5"].font = Font(bold=True)

            fila = 6
            def fila_concepto(nombre, valor, color=None):
                nonlocal fila
                ws[f"A{fila}"] = nombre
                ws[f"B{fila}"] = valor
                ws[f"B{fila}"].number_format = '"$"#,##0.00;-"$"#,##0.00'
                if color: ws[f"B{fila}"].font = Font(color=color, bold=True)
                fila += 1

            fila_concepto("Venta Excel", state.get("base_excel", 0))
            fila_concepto("Venta GNC", state.get("gnc_asignado", 0))
            fila_concepto("Total QR", state.get("qr_db", 0))
            fila_concepto("Productos Facturados", parse_moneda_robusto(state.get("prod", 0)))
            fila_concepto("Percepción", parse_moneda_robusto(state.get("per", 0)))
            fila_concepto("Tarjetas Ingresadas", parse_moneda_robusto(state.get("tarj", 0)))

            # Dinero Rendido (sum from rendiciones)
            movs = self.datos_rendiciones.get(v, {}).get("movimientos", [])
            efectivo = sum(
                parse_moneda_robusto(m["monto"])
                for m in movs if m.get("origen") in ("sql", "manual")
            )
            fila_concepto("Dinero Rendido", efectivo)

            fila_concepto("Total Anotaciones", self.total_anotaciones_por_vendedor(v))

            diff = balances.get(v, 0.0)
            fila_concepto("Diferencia", diff, "c0392b" if diff < 0 else "27ae60")

            fila += 1
            ws[f"A{fila}"] = "OBSERVACIONES"
            ws[f"A{fila}"].font = Font(bold=True)
            fila += 1

            textos = []
            for a in self.anotaciones_tmp:
                if son_nombres_similares(v, a.get("vendedor")):
                    desc = a.get("descripcion", "").strip()
                    if desc: textos.append(f"* {desc}")

            ws.merge_cells(start_row=fila, start_column=1, end_row=fila + max(3, len(textos)), end_column=2)
            ws[f"A{fila}"] = "\n".join(textos) if textos else "—"
            ws[f"A{fila}"].alignment = Alignment(vertical="top", wrap_text=True)
            ajustar_columnas(ws)

        # QR Detail
        ws_qr = wb.create_sheet("DETALLE_QR")
        aplicar_header(ws_qr, ["Planilla", "Operario", "Fecha", "Comprobante", "Estado", "Transacciones", "Importe Neto", "Extracash", "Descuentos", "Importe Completo"])

        for df in self.datos_detalle_qr.values():
            if df is None or df.empty: continue
            for _, r in df.iterrows():
                ws_qr.append([
                    self.planilla_desde_sql or "",
                    r.get("Operario"),
                    r.get("FEC"),
                    f"{r.get('TIP','')}-{r.get('TCO','')}-{r.get('NCO','')}",
                    r.get("Estado_QR"),
                    normalizar_id_transaccion(r.get("ID_TRANSACCION")),
                    r.get("IMPORTE"),
                    r.get("CASHOUT"),
                    r.get("DESC_PROMO"),
                    r.get("QR_FINAL")
                ])
        ajustar_columnas(ws_qr)

        # Product Detail
        ws_prod = wb.create_sheet("DETALLE_PRODUCTOS_FACTURADOS")
        if not self.datos_detalle_facturacion:
             ws_prod.append(["Sin productos facturados"])
        else:
             headers_prod = next((list(df.columns) for df in self.datos_detalle_facturacion.values() if df is not None and not df.empty), None)
             if headers_prod:
                 aplicar_header(ws_prod, headers_prod)
                 for df in self.datos_detalle_facturacion.values():
                     if df is None or df.empty: continue
                     for _, row in df.iterrows():
                         ws_prod.append(list(row.values))
        ajustar_columnas(ws_prod)

        wb.save(ruta)
        return ruta


    def unir_cajas(self, vendor_a, vendor_b):
        """
        Merges two vendors into the same 'caja'.
        Updates the caja_id for all vendors involved.
        """
        if vendor_a not in self.vendor_states or vendor_b not in self.vendor_states:
             raise ValueError("Vendedor no encontrado")

        state_a = self.vendor_states[vendor_a]
        state_b = self.vendor_states[vendor_b]

        # Create new caja_id (union of sets)
        new_caja_id = state_a["caja_id"] | state_b["caja_id"]

        # Update all vendors in this new group
        for v in new_caja_id:
            if v in self.vendor_states:
                self.vendor_states[v]["caja_id"] = new_caja_id

        # Mark B (and others except A?) as fusionado?
        # Original logic: B becomes fusionado.
        # But if A was already a group, and B was a group...
        # We need to decide who is the 'principal'.
        # Usually the one we are dragging TO (vendor_a) remains principal?
        # The UI logic was: "unir B a A", A keeps widgets enabled, B gets disabled.

        # Here we just mark fusionado.
        self.vendor_states[vendor_b]["fusionado"] = True
        # Note: In a pure logic class, we don't care about UI enabling/disabling.
        # Calculation logic will sum up everything for the unique caja_id.
